"""Importador de clientes y consumos desde el Excel histórico "Clientes .xlsx".

Cada hoja del Excel es un cliente. Las hojas contienen bloques de 6 columnas
(Cantidad | Producto | Dia | Precio | Total | separador) en orden cronológico
de izquierda a derecha; dentro de cada bloque los meses se apilan verticalmente
y se cierran con una fila-etiqueta (Producto = nombre del mes, Total = total).
El segmento final suele quedar sin etiquetar (mes en curso).

El script crea todos los clientes y luego importa los consumos de los meses
objetivo (abril-julio 2026) replicando la semántica del endpoint
`pedido_personalizado`: get-or-create de Producto por nombre y Transaccion con
precio_historico congelado. Cada mes queda como CuentaMensual "abierta".

Uso (desde la raíz del proyecto):
    backend/venv/bin/python -m backend.importar_excel --dry-run "/ruta/Clientes .xlsx"
    backend/venv/bin/python -m backend.importar_excel "/ruta/Clientes .xlsx"
"""

import argparse
import calendar
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

import openpyxl

from backend.base_datos import SesionLocal
from backend.modelos import Cliente, CuentaMensual, Producto, Transaccion

ANIO_OBJETIVO = 2026
MESES_OBJETIVO = (4, 5, 6, 7)
NOMBRE_MES = {4: "abril", 5: "mayo", 6: "junio", 7: "julio"}
# Hojas sin formato itemizado: solo se crea el cliente (carga manual posterior)
HOJAS_SOLO_CLIENTE = {"DANIL", "FRANCISCA"}

# Nombres largos primero para que "SEPTIEMBRE" no matchee antes como "SEPT"
_NOMBRES_MES = [
    ("DICIEMBRE", 12), ("NOVIEMBRE", 11), ("SEPTIEMBRE", 9), ("FEBRERO", 2),
    ("OCTUBRE", 10), ("AGOSTO", 8), ("ENERO", 1), ("MARZO", 3), ("ABRIL", 4),
    ("MAYO", 5), ("JUNIO", 6), ("JULIO", 7), ("SEPT", 9), ("FEB", 2),
    ("OCT", 10), ("NOV", 11), ("DIC", 12),
]


def _normalizar(texto: str) -> str:
    s = unicodedata.normalize("NFKD", texto.upper().strip())
    return "".join(c for c in s if not unicodedata.combining(c))


def detectar_meses_etiqueta(texto: str) -> list[int]:
    """Meses mencionados en una fila-etiqueta, en orden de aparición."""
    plano = _normalizar(texto)
    ocupado = [False] * len(plano)
    encontrados: list[tuple[int, int]] = []
    for nombre, mes in _NOMBRES_MES:
        for m in re.finditer(nombre, plano):
            if any(ocupado[m.start():m.end()]):
                continue
            for i in range(m.start(), m.end()):
                ocupado[i] = True
            encontrados.append((m.start(), mes))
    return [mes for _, mes in sorted(encontrados)]


def _vacio(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _a_decimal(v) -> Decimal | None:
    if _vacio(v):
        return None
    try:
        return Decimal(str(v).strip().replace(",", ".")).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _a_dia(v) -> int | None:
    if _vacio(v):
        return None
    if isinstance(v, datetime):
        return v.day
    try:
        d = int(float(str(v).strip()))
    except ValueError:
        return None
    return d if 1 <= d <= 31 else None


@dataclass
class Item:
    cantidad: float
    producto: str
    dia: int | None
    precio: Decimal


@dataclass
class Segmento:
    items: list[Item] = field(default_factory=list)
    meses_etiqueta: list[int] = field(default_factory=list)  # vacío = sin etiqueta
    total_etiqueta: Decimal | None = None
    anio: int | None = None  # asignados por fechar_segmentos
    mes: int | None = None


def parsear_hoja(ws, avisos: list[str]) -> list[Segmento]:
    """Secuencia cronológica de segmentos: bloques izq->der, filas arriba->abajo."""
    segmentos: list[Segmento] = []
    b = 0
    while b * 6 + 1 <= ws.max_column:
        c0 = b * 6 + 1
        actual = Segmento()
        for r in range(2, ws.max_row + 1):
            cant = ws.cell(row=r, column=c0).value
            prod = ws.cell(row=r, column=c0 + 1).value
            dia = ws.cell(row=r, column=c0 + 2).value
            prec = ws.cell(row=r, column=c0 + 3).value
            tot = ws.cell(row=r, column=c0 + 4).value

            if _vacio(cant) and isinstance(prod, str) and not _vacio(prod):
                meses = detectar_meses_etiqueta(prod)
                if meses:
                    actual.meses_etiqueta = meses
                    actual.total_etiqueta = _a_decimal(tot)
                    segmentos.append(actual)
                    actual = Segmento()
                    continue
                # Ítem con cantidad en blanco (cantidad implícita 1): se distingue de
                # una nota (abonos, comentarios) porque tiene día y precio numéricos
                precio = _a_decimal(prec) if not isinstance(prec, datetime) else None
                dia_num = _a_dia(dia)
                if precio is not None and dia_num is not None:
                    total = _a_decimal(tot)
                    cantidad = float(total / precio) if total and precio and total % precio == 0 else 1.0
                    actual.items.append(Item(cantidad, prod.strip(), dia_num, precio))
                else:
                    avisos.append(f"{ws.title}: nota ignorada fila {r}: {prod.strip()!r}")
                continue

            if not _vacio(cant) and not _vacio(prod):
                try:
                    cantidad = float(str(cant).strip())
                except ValueError:
                    avisos.append(f"{ws.title}: fila {r} con cantidad ilegible, omitida")
                    continue
                precio = _a_decimal(prec)
                total = _a_decimal(tot)
                if precio is None and total is not None and cantidad:
                    precio = (total / Decimal(str(cantidad))).quantize(Decimal("0.01"))
                if precio is None:
                    avisos.append(f"{ws.title}: fila {r} sin precio, omitida")
                    continue
                actual.items.append(Item(cantidad, str(prod).strip(), _a_dia(dia), precio))
                continue

            # Fila de cierre sin etiqueta: solo columna Total con valor
            if _vacio(cant) and _vacio(prod) and not _vacio(tot) and actual.items:
                actual.total_etiqueta = _a_decimal(tot)
                segmentos.append(actual)
                actual = Segmento()

        if actual.items:  # fin de bloque sin fila de cierre
            segmentos.append(actual)
        b += 1
    return segmentos


def _retroceder_mes(anio: int, mes: int) -> tuple[int, int]:
    return (anio - 1, 12) if mes == 1 else (anio, mes - 1)


def _avanzar_mes(anio: int, mes: int) -> tuple[int, int]:
    return (anio + 1, 1) if mes == 12 else (anio, mes + 1)


def fechar_segmentos(segmentos: list[Segmento]) -> None:
    """Asigna (anio, mes) a cada segmento caminando desde el ancla final.

    El último segmento etiquetado ancla el año: si su mes es <= junio pertenece
    al año objetivo (los datos más a la derecha son "de este año"); si es julio
    o posterior, es del año anterior. Hacia atrás cada etiqueta toma la última
    ocurrencia posible de su mes; los segmentos sin etiqueta toman el mes
    anterior. Hacia adelante (cola sin etiquetar) se avanza mes a mes.
    """
    etiquetados = [i for i, s in enumerate(segmentos) if s.meses_etiqueta]
    if not etiquetados:
        return
    a = etiquetados[-1]
    mes_ancla = segmentos[a].meses_etiqueta[-1]
    anio_ancla = ANIO_OBJETIVO if mes_ancla <= max(MESES_OBJETIVO) else ANIO_OBJETIVO - 1
    segmentos[a].anio, segmentos[a].mes = anio_ancla, mes_ancla

    # Hacia atrás
    anio, mes = anio_ancla, segmentos[a].meses_etiqueta[0]
    if segmentos[a].meses_etiqueta[0] > mes_ancla:  # etiqueta que cruza año (DIC&ENE)
        anio -= 1
    for s in reversed(segmentos[:a]):
        if s.meses_etiqueta:
            ultimo = s.meses_etiqueta[-1]
            s.anio = anio if ultimo < mes else anio - 1
            s.mes = ultimo
            anio, mes = s.anio, s.meses_etiqueta[0]
            if s.meses_etiqueta[0] > ultimo:
                anio -= 1
        else:
            anio, mes = _retroceder_mes(anio, mes)
            s.anio, s.mes = anio, mes

    # Hacia adelante (cola en curso, sin etiqueta)
    anio, mes = anio_ancla, mes_ancla
    for s in segmentos[a + 1:]:
        anio, mes = _avanzar_mes(anio, mes)
        s.anio, s.mes = anio, mes


@dataclass
class MesImportable:
    mes: int
    items: list[Item]
    total_referencia: Decimal | None  # total de la fila-etiqueta del Excel, si aplica
    inferido: bool  # True si el mes se dedujo (segmento sin etiqueta o dividido)

    @property
    def total_calculado(self) -> Decimal:
        return sum((Decimal(str(i.cantidad)) * i.precio for i in self.items),
                   Decimal("0.00"))


def extraer_meses_objetivo(segmentos: list[Segmento], avisos: list[str],
                           hoja: str) -> dict[int, MesImportable]:
    """Devuelve {mes: MesImportable} para (ANIO_OBJETIVO, MESES_OBJETIVO)."""
    resultado: dict[int, MesImportable] = {}

    def acumular(anio: int, mes: int, items: list[Item],
                 ref: Decimal | None, inferido: bool):
        if anio != ANIO_OBJETIVO or mes not in MESES_OBJETIVO or not items:
            return
        if mes in resultado:
            resultado[mes].items.extend(items)
            resultado[mes].total_referencia = None  # ya no es comparable 1:1
            resultado[mes].inferido = True
            avisos.append(f"{hoja}: {NOMBRE_MES[mes]} aparece en más de un segmento; fusionado")
        else:
            resultado[mes] = MesImportable(mes, list(items), ref, inferido)

    for s in segmentos:
        if s.anio is None or s.mes is None or not s.items:
            continue
        # Nota: los segmentos suelen incluir al inicio ítems del día 30/31 del mes
        # anterior (facturados en este mes); un salto de días hacia el final puede
        # indicar ítems del mes siguiente aún sin cerrar — solo se avisa.
        dias = [i.dia for i in s.items if i.dia is not None]
        for idx, (a, b) in enumerate(zip(dias, dias[1:])):
            if idx >= 3 and b < a - 7:
                avisos.append(f"{hoja}: salto de día {a}->{b} al final del segmento de "
                              f"{s.mes}/{s.anio}; revisar si los últimos ítems son del mes siguiente")
        acumular(s.anio, s.mes, s.items, s.total_etiqueta, not s.meses_etiqueta)
    return resultado


def limpiar_nombre_producto(nombre: str, avisos: list[str], hoja: str) -> str:
    limpio = re.sub(r"\s+", " ", nombre).strip()
    if "<" in limpio or ">" in limpio:
        limpio = limpio.replace("<", "").replace(">", "")
        avisos.append(f"{hoja}: producto con caracteres HTML saneado: {nombre!r}")
    return limpio[:100]


@dataclass
class ResumenHoja:
    cliente: str
    meses: dict[int, MesImportable]
    motivo_sin_datos: str | None = None


def analizar_libro(ruta: str) -> tuple[list[ResumenHoja], list[str]]:
    avisos: list[str] = []
    resumenes: list[ResumenHoja] = []
    libro = openpyxl.load_workbook(ruta, data_only=True)
    for nombre_hoja in libro.sheetnames:
        ws = libro[nombre_hoja]
        cliente = re.sub(r"\s+", " ", nombre_hoja).strip().title()
        if _normalizar(nombre_hoja) in HOJAS_SOLO_CLIENTE:
            resumenes.append(ResumenHoja(cliente, {}, "formato no estándar (carga manual)"))
            continue
        segmentos = parsear_hoja(ws, avisos)
        if not any(s.items for s in segmentos):
            resumenes.append(ResumenHoja(cliente, {}, "hoja sin consumos"))
            continue
        fechar_segmentos(segmentos)
        if segmentos and segmentos[0].anio is None:
            resumenes.append(ResumenHoja(cliente, {}, "sin etiquetas de mes (no fechable)"))
            continue
        meses = extraer_meses_objetivo(segmentos, avisos, cliente)
        if not meses:
            ult = next((s for s in reversed(segmentos) if s.items), None)
            detalle = f"últimos datos: {ult.mes}/{ult.anio}" if ult else "sin datos"
            resumenes.append(ResumenHoja(cliente, {}, f"sin consumos abr-jun {ANIO_OBJETIVO} ({detalle})"))
            continue
        resumenes.append(ResumenHoja(cliente, meses))
    return resumenes, avisos


def importar(resumenes: list[ResumenHoja], avisos: list[str]) -> dict[str, int]:
    contadores = {"clientes_creados": 0, "cuentas_creadas": 0, "transacciones_creadas": 0,
                  "productos_creados": 0, "meses_omitidos": 0}
    db = SesionLocal()
    productos_de_esta_corrida: set[str] = set()
    try:
        for res in resumenes:
            try:
                cliente = db.query(Cliente).filter(Cliente.nombre == res.cliente).first()
                if not cliente:
                    cliente = Cliente(nombre=res.cliente, estado="activo")
                    db.add(cliente)
                    db.flush()
                    contadores["clientes_creados"] += 1

                for mes in sorted(res.meses):
                    datos = res.meses[mes]
                    existente = db.query(CuentaMensual).filter(
                        CuentaMensual.cliente_id == cliente.id,
                        CuentaMensual.mes == mes,
                        CuentaMensual.anio == ANIO_OBJETIVO,
                    ).first()
                    if existente and existente.transacciones:
                        avisos.append(f"{res.cliente}: {NOMBRE_MES[mes]} ya tiene datos en la BD, omitido")
                        contadores["meses_omitidos"] += 1
                        continue
                    cuenta = existente or CuentaMensual(
                        cliente_id=cliente.id, mes=mes, anio=ANIO_OBJETIVO,
                        porcentaje_descuento=Decimal("0.00"), estado="abierta",
                    )
                    if not existente:
                        db.add(cuenta)
                        db.flush()
                        contadores["cuentas_creadas"] += 1

                    for item in datos.items:
                        nombre_prod = limpiar_nombre_producto(item.producto, avisos, res.cliente)
                        producto = db.query(Producto).filter(Producto.nombre == nombre_prod).first()
                        if not producto:
                            producto = Producto(nombre=nombre_prod, precio_actual=item.precio,
                                                stock_actual=0)
                            db.add(producto)
                            db.flush()
                            contadores["productos_creados"] += 1
                            productos_de_esta_corrida.add(nombre_prod)
                        elif nombre_prod in productos_de_esta_corrida:
                            # Mantener como precio_actual el último visto en la importación,
                            # sin pisar precios de productos que ya existían en la BD
                            producto.precio_actual = item.precio

                        cantidad = item.cantidad
                        precio = item.precio
                        if cantidad != int(cantidad) or cantidad < 1:
                            precio = (Decimal(str(cantidad)) * precio).quantize(Decimal("0.01"))
                            avisos.append(f"{res.cliente}: cantidad {cantidad} no entera en "
                                          f"{nombre_prod!r}, importada como 1 x {precio}")
                            cantidad = 1
                        dia = min(item.dia or 15, calendar.monthrange(ANIO_OBJETIVO, mes)[1])
                        db.add(Transaccion(
                            cuenta_mensual_id=cuenta.id,
                            producto_id=producto.id,
                            cantidad=int(cantidad),
                            precio_historico=precio,
                            fecha_hora=datetime(ANIO_OBJETIVO, mes, dia, 12, 0,
                                                tzinfo=timezone.utc),
                        ))
                        contadores["transacciones_creadas"] += 1
                db.commit()
            except Exception as exc:  # noqa: BLE001 - reportar y seguir con la siguiente hoja
                db.rollback()
                avisos.append(f"ERROR en {res.cliente}: {exc!r} (hoja revertida)")
    finally:
        db.close()
    return contadores


def formatear_reporte(resumenes: list[ResumenHoja], avisos: list[str]) -> str:
    lineas = ["", "=== Clientes con consumos abril-junio a importar ==="]
    total_general = Decimal("0.00")
    for res in resumenes:
        if not res.meses:
            continue
        partes = []
        for mes in sorted(res.meses):
            d = res.meses[mes]
            calc = d.total_calculado
            total_general += calc
            if d.total_referencia is not None:
                marca = "OK" if calc == d.total_referencia else f"Excel={d.total_referencia:,.0f} DIFIERE"
            else:
                marca = "mes inferido" if d.inferido else "s/ref"
            partes.append(f"{NOMBRE_MES[mes]}: {len(d.items)} items ${calc:,.0f} [{marca}]")
        lineas.append(f"  {res.cliente:<15} " + " | ".join(partes))
    lineas.append(f"  TOTAL GENERAL: ${total_general:,.0f}")

    lineas.append("\n=== Clientes que se crean sin consumos ===")
    for res in resumenes:
        if res.motivo_sin_datos:
            lineas.append(f"  {res.cliente:<15} {res.motivo_sin_datos}")

    if avisos:
        lineas.append("\n=== Avisos ===")
        lineas.extend(f"  {a}" for a in avisos)
    return "\n".join(lineas)


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa clientes y consumos desde el Excel")
    parser.add_argument("ruta", help="Ruta del archivo .xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo analizar y reportar, sin escribir en la BD")
    args = parser.parse_args()

    resumenes, avisos = analizar_libro(args.ruta)
    if args.dry_run:
        print(formatear_reporte(resumenes, avisos))
        print("\n(dry-run: no se escribió nada en la base de datos)")
        return

    contadores = importar(resumenes, avisos)
    print(formatear_reporte(resumenes, avisos))
    print("\n=== Resultado ===")
    for k, v in contadores.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
