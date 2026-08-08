"""Generación del libro Excel de balances mensuales.

Reutiliza los cálculos de ``backend/routers/balances.py`` para que el archivo y
la pantalla muestren exactamente los mismos números, y agrega dos consultas de
detalle (cuentas y transacciones) que solo tienen sentido en el Excel.
"""
import datetime
import io
from decimal import Decimal
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from backend.constantes import NOMBRES_MESES, PRODUCTOS_ARRASTRE
from backend.modelos import Categoria, Cliente, CuentaMensual, PerdidaInventario, Producto, Transaccion
from backend.routers.balances import _dec, _q, _rango_utc, _sub_totales_cuenta, construir_balance

FORMATO_CLP = '"$"#,##0;[Red]-"$"#,##0'
FORMATO_ENTERO = "#,##0"
FORMATO_PCT = '0.0"%"'
FORMATO_FECHA = "dd-mm-yyyy hh:mm"

AZUL = "1E3A5F"
FILL_HEADER = PatternFill("solid", fgColor=AZUL)
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_TITULO = Font(bold=True, size=14, color=AZUL)
FONT_SUBTITULO = Font(italic=True, size=9, color="666666")
FONT_SECCION = Font(bold=True, size=11, color=AZUL)
FONT_TOTAL = Font(bold=True)
FONT_VACIO = Font(italic=True, color="888888")
BORDE_TOTAL = Border(top=Side(style="thin", color=AZUL))

MAX_FILAS_DETALLE = 20000 # tope defensivo; un mes real ronda las 900 filas


def _titulo(ws: Worksheet, texto: str, subtitulo: Optional[str] = None) -> int:
    """Escribe el bloque de título. Devuelve la fila donde sigue el contenido."""
    ws["A1"] = texto
    ws["A1"].font = FONT_TITULO
    if subtitulo:
        ws["A2"] = subtitulo
        ws["A2"].font = FONT_SUBTITULO
        return 4
    return 3


def _encabezados(ws: Worksheet, fila: int, columnas: List[str]) -> None:
    for i, titulo in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=i, value=titulo)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[fila].height = 22
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def _sin_datos(ws: Worksheet, fila: int, columnas: int, mensaje: str = "Sin registros en este período") -> None:
    celda = ws.cell(row=fila, column=1, value=mensaje)
    celda.font = FONT_VACIO
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=max(columnas, 1))


def _autoajustar(ws: Worksheet, minimo: int = 10, maximo: int = 45) -> None:
    anchos = {}
    for fila in ws.iter_rows():
        for celda in fila:
            if celda.value is None:
                continue
            largo = len(str(celda.value))
            anchos[celda.column] = max(anchos.get(celda.column, 0), largo)
    for col, largo in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = max(minimo, min(maximo, largo + 3))


def _escribir_filas(ws: Worksheet, fila_inicio: int, filas: List[list], formatos: dict) -> int:
    """Escribe las filas aplicando formato por columna (1-indexed). Devuelve la
    siguiente fila libre."""
    fila = fila_inicio
    for datos in filas:
        for i, valor in enumerate(datos, start=1):
            celda = ws.cell(row=fila, column=i, value=valor)
            if i in formatos:
                celda.number_format = formatos[i]
        fila += 1
    return fila


def _fila_total(ws: Worksheet, fila: int, etiqueta: str, valores: dict, formatos: dict) -> None:
    celda = ws.cell(row=fila, column=1, value=etiqueta)
    celda.font = FONT_TOTAL
    celda.border = BORDE_TOTAL
    max_col = max(list(valores.keys()) + list(formatos.keys()) + [1])
    for col in range(2, max_col + 1):
        c = ws.cell(row=fila, column=col, value=valores.get(col))
        c.font = FONT_TOTAL
        c.border = BORDE_TOTAL
        if col in formatos:
            c.number_format = formatos[col]


# --- Hojas -------------------------------------------------------------------------

def _hoja_resumen(wb: Workbook, balance, mes: int, anio: int) -> None:
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_properties.tabColor = AZUL
    ws.sheet_view.showGridLines = False
    generado = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
    fila = _titulo(ws, f"Balance de {NOMBRES_MESES[mes - 1]} {anio}", f"Generado el {generado}")

    r = balance.resumen
    mes_ant = NOMBRES_MESES[(mes - 2) % 12]
    _encabezados(ws, fila, ["Indicador", "Mes actual", f"{mes_ant} (anterior)", "Variación"])
    fila += 1

    def bloque(nombre: str, filas_bloque: list) -> None:
        nonlocal fila
        celda = ws.cell(row=fila, column=1, value=nombre)
        celda.font = FONT_SECCION
        fila += 1
        for etiqueta, kpi, formato in filas_bloque:
            ws.cell(row=fila, column=1, value=etiqueta)
            c_act = ws.cell(row=fila, column=2, value=float(kpi.actual))
            c_ant = ws.cell(row=fila, column=3, value=float(kpi.anterior))
            c_act.number_format = formato
            c_ant.number_format = formato
            c_var = ws.cell(row=fila, column=4,
                            value=round(kpi.variacion_pct, 1) if kpi.variacion_pct is not None else "—")
            if kpi.variacion_pct is not None:
                c_var.number_format = FORMATO_PCT
            fila += 1
        fila += 1

    bloque("Resultado del mes", [
        ("Ventas (consumo real)", r.ventas, FORMATO_CLP),
        ("Cobrado", r.cobrado, FORMATO_CLP),
        ("Por cobrar", r.por_cobrar, FORMATO_CLP),
        ("Descuentos otorgados", r.descuentos, FORMATO_CLP),
        ("Ticket promedio por cliente", r.ticket_promedio, FORMATO_CLP),
    ])
    bloque("Actividad", [
        ("Clientes activos", r.clientes_activos, FORMATO_ENTERO),
        ("Unidades vendidas", r.unidades_vendidas, FORMATO_ENTERO),
    ])

    # Escalares sin comparación
    for etiqueta, valor, formato in [
        ("Cuentas abiertas", r.cuentas_abiertas, FORMATO_ENTERO),
        ("Cuentas pagadas", r.cuentas_pagadas, FORMATO_ENTERO),
        ("Tasa de cobro", round(r.tasa_cobro_pct, 1) if r.tasa_cobro_pct is not None else "—",
         FORMATO_PCT if r.tasa_cobro_pct is not None else None),
    ]:
        ws.cell(row=fila, column=1, value=etiqueta)
        c = ws.cell(row=fila, column=2, value=valor)
        if formato:
            c.number_format = formato
        fila += 1
    fila += 1

    ws.cell(row=fila, column=1, value="Inventario").font = FONT_SECCION
    fila += 1
    ws.cell(row=fila, column=1, value="Valor de mermas (a precio de venta)")
    c = ws.cell(row=fila, column=2, value=float(r.valor_mermas.actual))
    c.number_format = FORMATO_CLP
    c_ant = ws.cell(row=fila, column=3, value=float(r.valor_mermas.anterior))
    c_ant.number_format = FORMATO_CLP
    fila += 1
    ws.cell(row=fila, column=1, value="Productos con stock bajo")
    ws.cell(row=fila, column=2, value=len(balance.stock_bajo)).number_format = FORMATO_ENTERO
    fila += 2

    ws.cell(row=fila, column=1, value="Conciliación de deuda").font = FONT_SECCION
    fila += 1
    for etiqueta, valor in [
        ("Deuda recibida del mes anterior", balance.resumen.deuda_arrastrada),
        ("Deuda traspasada al mes siguiente", balance.resumen.deuda_traspasada),
        ("Clientes con deuda pendiente", Decimal(balance.resumen.clientes_con_deuda)),
    ]:
        ws.cell(row=fila, column=1, value=etiqueta)
        c = ws.cell(row=fila, column=2, value=float(valor))
        c.number_format = FORMATO_CLP if "Deuda" in etiqueta else FORMATO_ENTERO
        fila += 1

    fila += 1
    nota = ws.cell(row=fila, column=1, value=(
        "Nota: 'Cobrado' y 'Por cobrar' incluyen los movimientos de arrastre de deuda; "
        "'Ventas' corresponde solo al consumo real del mes."
    ))
    nota.font = FONT_SUBTITULO
    _autoajustar(ws, minimo=14, maximo=48)


def _hoja_deudores(wb: Workbook, balance) -> None:
    ws = wb.create_sheet("Deudores")
    ws.sheet_properties.tabColor = "C0392B"
    fila = _titulo(ws, "Deudores", "Deuda del mes seleccionado y deuda total acumulada (todas las cuentas abiertas)")
    columnas = ["Cliente", "Teléfono", "Deuda del mes", "Deuda total acumulada", "Cuentas abiertas"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    if not balance.deudores:
        _sin_datos(ws, inicio, len(columnas), "Nadie tiene deuda pendiente en este período")
    else:
        filas = [
            [d.nombre, d.telefono or "", float(d.deuda_mes), float(d.deuda_total), d.cuentas_abiertas]
            for d in balance.deudores
        ]
        siguiente = _escribir_filas(ws, inicio, filas,
                                    {3: FORMATO_CLP, 4: FORMATO_CLP, 5: FORMATO_ENTERO})
        _fila_total(ws, siguiente, "TOTAL", {
            3: float(sum(d.deuda_mes for d in balance.deudores)),
            4: float(sum(d.deuda_total for d in balance.deudores)),
        }, {3: FORMATO_CLP, 4: FORMATO_CLP})
        ws.auto_filter.ref = f"A{fila}:E{siguiente - 1}"
    _autoajustar(ws)


def _hoja_cuentas(wb: Workbook, db: Session, mes: int, anio: int) -> None:
    ws = wb.create_sheet("Cuentas del mes")
    fila = _titulo(ws, "Cuentas del mes", "Una fila por cuenta (un cliente puede tener más de una en el mismo mes)")
    columnas = ["Cliente", "Consumo bruto", "Descuento %", "Descuento $", "Total cuenta", "Estado"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    tot = _sub_totales_cuenta(db)
    registros = (
        db.query(
            Cliente.nombre,
            func.coalesce(tot.c.bruto, 0),
            CuentaMensual.porcentaje_descuento,
            CuentaMensual.estado,
        )
        .join(CuentaMensual, CuentaMensual.cliente_id == Cliente.id)
        .outerjoin(tot, tot.c.cuenta_id == CuentaMensual.id)
        .filter(CuentaMensual.mes == mes, CuentaMensual.anio == anio, tot.c.bruto.isnot(None))
        .order_by(Cliente.nombre)
        .all()
    )
    if not registros:
        _sin_datos(ws, inicio, len(columnas))
    else:
        filas, total_bruto, total_neto = [], Decimal("0"), Decimal("0")
        for nombre, bruto, pct, estado in registros:
            bruto_d, pct_d = _dec(bruto), _dec(pct)
            desc = bruto_d * pct_d / 100
            neto = bruto_d - desc
            total_bruto += bruto_d
            total_neto += neto
            filas.append([nombre, float(bruto_d), float(pct_d), float(desc), float(neto),
                          "Pagada" if estado == "pagada" else "Abierta"])
        siguiente = _escribir_filas(ws, inicio, filas,
                                    {2: FORMATO_CLP, 3: '0.0"%"', 4: FORMATO_CLP, 5: FORMATO_CLP})
        _fila_total(ws, siguiente, "TOTAL",
                    {2: float(total_bruto), 5: float(total_neto)},
                    {2: FORMATO_CLP, 5: FORMATO_CLP})
        ws.auto_filter.ref = f"A{fila}:F{siguiente - 1}"
    _autoajustar(ws)


def _hoja_productos(wb: Workbook, db: Session, mes: int, anio: int) -> None:
    ws = wb.create_sheet("Productos")
    fila = _titulo(ws, "Ventas por producto", "Excluye los movimientos de arrastre de deuda")
    columnas = ["Producto", "Categoría", "Unidades", "Monto vendido", "% del total"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    registros = (
        db.query(
            Producto.nombre,
            Categoria.nombre,
            func.sum(Transaccion.cantidad),
            func.sum(Transaccion.cantidad * Transaccion.precio_historico),
        )
        .join(Transaccion, Transaccion.producto_id == Producto.id)
        .join(CuentaMensual, CuentaMensual.id == Transaccion.cuenta_mensual_id)
        .outerjoin(Categoria, Categoria.id == Producto.categoria_id)
        .filter(CuentaMensual.mes == mes, CuentaMensual.anio == anio,
                Producto.nombre.notin_(PRODUCTOS_ARRASTRE))
        .group_by(Producto.nombre, Categoria.nombre)
        .order_by(func.sum(Transaccion.cantidad * Transaccion.precio_historico).desc())
        .all()
    )
    if not registros:
        _sin_datos(ws, inicio, len(columnas))
    else:
        total = sum(_dec(m) for _n, _c, _u, m in registros) or Decimal("1")
        filas = [
            [n, c or "Sin categoría", int(u or 0), float(_dec(m)), float(_dec(m) / total * 100)]
            for n, c, u, m in registros
        ]
        siguiente = _escribir_filas(ws, inicio, filas,
                                    {3: FORMATO_ENTERO, 4: FORMATO_CLP, 5: FORMATO_PCT})
        _fila_total(ws, siguiente, "TOTAL", {
            3: sum(int(u or 0) for _n, _c, u, _m in registros),
            4: float(total),
        }, {3: FORMATO_ENTERO, 4: FORMATO_CLP})
        ws.auto_filter.ref = f"A{fila}:E{siguiente - 1}"
    _autoajustar(ws)


def _hoja_categorias(wb: Workbook, balance) -> None:
    ws = wb.create_sheet("Categorías")
    fila = _titulo(ws, "Ventas por categoría")
    columnas = ["Categoría", "Unidades", "Monto", "% del total"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    if not balance.ventas_por_categoria:
        _sin_datos(ws, inicio, len(columnas))
    else:
        total = sum(c.monto for c in balance.ventas_por_categoria) or Decimal("1")
        filas = [
            [c.nombre, c.unidades, float(c.monto), float(c.monto / total * 100)]
            for c in balance.ventas_por_categoria
        ]
        siguiente = _escribir_filas(ws, inicio, filas,
                                    {2: FORMATO_ENTERO, 3: FORMATO_CLP, 4: FORMATO_PCT})
        _fila_total(ws, siguiente, "TOTAL", {
            2: sum(c.unidades for c in balance.ventas_por_categoria),
            3: float(total),
        }, {2: FORMATO_ENTERO, 3: FORMATO_CLP})
    _autoajustar(ws)


def _hoja_mermas(wb: Workbook, db: Session, mes: int, anio: int) -> None:
    ws = wb.create_sheet("Mermas")
    ws.sheet_properties.tabColor = "E67E22"
    fila = _titulo(ws, "Mermas del mes",
                   "Los valores están a PRECIO DE VENTA: el sistema no registra costo de compra.")
    columnas = ["Fecha", "Producto", "Categoría", "Cantidad", "Valor unitario", "Valor total", "Motivo"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    ini, fin = _rango_utc(mes, anio)
    registros = (
        db.query(PerdidaInventario)
        .options(joinedload(PerdidaInventario.producto).joinedload(Producto.categoria))
        .filter(PerdidaInventario.fecha_hora >= ini, PerdidaInventario.fecha_hora < fin)
        .order_by(PerdidaInventario.fecha_hora)
        .all()
    )
    if not registros:
        _sin_datos(ws, inicio, len(columnas), "Sin mermas registradas en este período")
    else:
        filas, total = [], Decimal("0")
        for p in registros:
            valor = _dec(p.cantidad) * _dec(p.costo_historico)
            total += valor
            filas.append([
                p.fecha_hora.replace(tzinfo=None) if p.fecha_hora else None,
                p.producto.nombre if p.producto else "(producto eliminado)",
                (p.producto.categoria.nombre if p.producto and p.producto.categoria else "Sin categoría"),
                p.cantidad, float(_dec(p.costo_historico)), float(valor), p.motivo or "Sin motivo",
            ])
        siguiente = _escribir_filas(ws, inicio, filas, {
            1: FORMATO_FECHA, 4: FORMATO_ENTERO, 5: FORMATO_CLP, 6: FORMATO_CLP,
        })
        _fila_total(ws, siguiente, "TOTAL", {6: float(total)}, {6: FORMATO_CLP})
        ws.auto_filter.ref = f"A{fila}:G{siguiente - 1}"
    _autoajustar(ws)


def _hoja_detalle(wb: Workbook, db: Session, mes: int, anio: int) -> None:
    ws = wb.create_sheet("Detalle de consumos")
    fila = _titulo(ws, "Detalle de consumos",
                   "La columna 'Tipo' distingue las ventas reales de los movimientos de arrastre de deuda.")
    columnas = ["Fecha", "Cliente", "Producto", "Categoría", "Cantidad",
                "Precio unitario", "Subtotal", "Estado cuenta", "Tipo"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    registros = (
        db.query(Transaccion, Cliente.nombre, CuentaMensual.estado)
        .join(CuentaMensual, CuentaMensual.id == Transaccion.cuenta_mensual_id)
        .join(Cliente, Cliente.id == CuentaMensual.cliente_id)
        .options(joinedload(Transaccion.producto).joinedload(Producto.categoria))
        .filter(CuentaMensual.mes == mes, CuentaMensual.anio == anio)
        .order_by(Transaccion.fecha_hora)
        .limit(MAX_FILAS_DETALLE + 1)
        .all()
    )
    truncado = len(registros) > MAX_FILAS_DETALLE
    registros = registros[:MAX_FILAS_DETALLE]

    if not registros:
        _sin_datos(ws, inicio, len(columnas))
    else:
        filas, total = [], Decimal("0")
        for t, cliente, estado in registros:
            nombre_prod = t.producto.nombre if t.producto else "(pedido personalizado)"
            es_arrastre = nombre_prod in PRODUCTOS_ARRASTRE
            subtotal = _dec(t.cantidad) * _dec(t.precio_historico)
            total += subtotal
            filas.append([
                t.fecha_hora.replace(tzinfo=None) if t.fecha_hora else None,
                cliente, nombre_prod,
                (t.producto.categoria.nombre if t.producto and t.producto.categoria else "Sin categoría"),
                t.cantidad, float(_dec(t.precio_historico)), float(subtotal),
                "Pagada" if estado == "pagada" else "Abierta",
                "Arrastre de deuda" if es_arrastre else "Venta",
            ])
        siguiente = _escribir_filas(ws, inicio, filas, {
            1: FORMATO_FECHA, 5: FORMATO_ENTERO, 6: FORMATO_CLP, 7: FORMATO_CLP,
        })
        _fila_total(ws, siguiente, "TOTAL", {7: float(total)}, {7: FORMATO_CLP})
        if truncado:
            ws.cell(row=siguiente + 1, column=1,
                    value=f"… detalle truncado a {MAX_FILAS_DETALLE} filas").font = FONT_VACIO
        ws.auto_filter.ref = f"A{fila}:I{siguiente - 1}"
    _autoajustar(ws)


def _hoja_stock(wb: Workbook, balance) -> None:
    ws = wb.create_sheet("Alertas de stock")
    fila = _titulo(ws, "Alertas de stock",
                   "Foto del inventario al momento de generar el archivo (no depende del mes seleccionado).")
    columnas = ["Producto", "Categoría", "Stock actual", "Precio actual", "Valor en stock"]
    _encabezados(ws, fila, columnas)
    inicio = fila + 1

    if not balance.stock_bajo:
        _sin_datos(ws, inicio, len(columnas), "Todo el stock está sobre el umbral")
    else:
        filas = [
            [p.nombre, p.categoria or "Sin categoría", p.stock_actual,
             float(p.precio_actual), float(p.precio_actual * p.stock_actual)]
            for p in balance.stock_bajo
        ]
        _escribir_filas(ws, inicio, filas, {3: FORMATO_ENTERO, 4: FORMATO_CLP, 5: FORMATO_CLP})
    _autoajustar(ws)


def construir_libro_balances(db: Session, mes: int, anio: int) -> bytes:
    """Arma el libro completo del mes y lo devuelve como bytes."""
    balance = construir_balance(db, mes, anio)

    wb = Workbook()
    _hoja_resumen(wb, balance, mes, anio)
    _hoja_deudores(wb, balance)
    _hoja_cuentas(wb, db, mes, anio)
    _hoja_productos(wb, db, mes, anio)
    _hoja_categorias(wb, balance)
    _hoja_mermas(wb, db, mes, anio)
    _hoja_detalle(wb, db, mes, anio)
    _hoja_stock(wb, balance)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo(mes: int, anio: int) -> str:
    return f"Balance_{anio}-{mes:02d}_{NOMBRES_MESES[mes - 1]}.xlsx"
