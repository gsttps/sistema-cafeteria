"""Tests de las métricas de Balances, del rollover de deuda y del export a Excel."""
import io

import openpyxl
import pytest

HOJAS_ESPERADAS = [
    "Resumen", "Deudores", "Cuentas del mes", "Productos",
    "Categorías", "Mermas", "Detalle de consumos", "Alertas de stock",
]


# --- Helpers -----------------------------------------------------------------------

def _crear_cliente(client, nombre):
    resp = client.post("/api/v1/clientes/", json={"nombre": nombre})
    assert resp.status_code == 201
    return resp.json()


def _crear_producto(client, nombre, precio="1000", stock=100):
    resp = client.post(
        "/api/v1/productos/",
        json={"nombre": nombre, "precio_actual": precio, "stock_actual": stock},
    )
    assert resp.status_code == 201
    return resp.json()


def _agregar_item(client, cliente_id, producto_id, cantidad=1, mes=None, anio=None, dia=None):
    cuerpo = {"producto_id": producto_id, "cantidad": cantidad}
    if mes:
        cuerpo.update({"mes": mes, "anio": anio})
    if dia:
        cuerpo["dia"] = dia
    resp = client.post(f"/api/v1/cuentas/cliente/{cliente_id}/agregar_item", json=cuerpo)
    assert resp.status_code == 200
    return resp.json()


def _cuenta(client, cliente_id, mes=None, anio=None):
    params = {"mes": mes, "anio": anio} if mes else {}
    resp = client.get(f"/api/v1/cuentas/cliente/{cliente_id}", params=params)
    assert resp.status_code == 200
    return resp.json()


def _cerrar(client, cuenta_id, monto=None):
    resp = client.put(f"/api/v1/cuentas/{cuenta_id}/cerrar",
                      json={"monto_pagado": monto} if monto is not None else {})
    assert resp.status_code == 200
    return resp.json()


def _balance(client, mes, anio):
    resp = client.get("/api/v1/balances/", params={"mes": mes, "anio": anio})
    assert resp.status_code == 200
    return resp.json()


def _mes_actual(client):
    """Mes/año que usa el backend por defecto (los items sin mes van ahí)."""
    cliente = _crear_cliente(client, "Sonda Mes Actual")
    c = _cuenta(client, cliente["id"])
    return c["mes"], c["anio"]


# --- Métricas base -----------------------------------------------------------------

class TestMetricasBasicas:
    def test_mes_sin_datos_devuelve_ceros_sin_error(self, client, auth_headers):
        b = _balance(client, 3, 2024)
        r = b["resumen"]
        assert r["ventas"]["actual"] == 0
        assert r["cobrado"]["actual"] == 0
        assert r["por_cobrar"]["actual"] == 0
        assert r["ticket_promedio"]["actual"] == 0
        assert r["clientes_activos"]["actual"] == 0
        assert r["ventas"]["variacion_pct"] is None
        assert r["tasa_cobro_pct"] is None
        assert b["productos_top"] == [] and b["deudores"] == []

    def test_ventas_y_unidades_exactas(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Ventas")
        p1 = _crear_producto(client, "Café Ventas", "1500")
        p2 = _crear_producto(client, "Torta Ventas", "3000")
        _agregar_item(client, cliente["id"], p1["id"], cantidad=2)   # 3000
        _agregar_item(client, cliente["id"], p2["id"], cantidad=1)   # 3000

        r = _balance(client, mes, anio)["resumen"]
        assert r["ventas"]["actual"] == 6000
        assert r["unidades_vendidas"]["actual"] == 3
        assert r["clientes_activos"]["actual"] == 1

    def test_cuenta_pagada_va_a_cobrado(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Pagado")
        prod = _crear_producto(client, "Producto Pagado", "2000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=2)
        _cerrar(client, _cuenta(client, cliente["id"])["id"])

        r = _balance(client, mes, anio)["resumen"]
        assert r["cobrado"]["actual"] == 4000
        assert r["por_cobrar"]["actual"] == 0
        assert r["cuentas_pagadas"] == 1
        assert r["tasa_cobro_pct"] == 100

    def test_descuento_afecta_cobrado_y_descuentos(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Descuento")
        prod = _crear_producto(client, "Producto Descuento", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=10)  # 10.000
        cuenta = _cuenta(client, cliente["id"])
        client.put(f"/api/v1/cuentas/{cuenta['id']}/descuento", params={"porcentaje_descuento": 10})
        _cerrar(client, cuenta["id"])

        r = _balance(client, mes, anio)["resumen"]
        assert r["descuentos"]["actual"] == 1000
        assert r["cobrado"]["actual"] == 9000

    def test_ticket_promedio_divide_por_clientes_con_consumo(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        prod = _crear_producto(client, "Producto Ticket", "1000")
        for i in range(3):
            c = _crear_cliente(client, f"Cliente Ticket {i}")
            _agregar_item(client, c["id"], prod["id"], cantidad=3)  # 3000 c/u

        r = _balance(client, mes, anio)["resumen"]
        assert r["ventas"]["actual"] == 9000
        assert r["clientes_activos"]["actual"] == 3
        assert r["ticket_promedio"]["actual"] == 3000


class TestCuentasFantasma:
    """Consultar la cuenta de un cliente no debe crear nada, y las cuentas que
    igual quedan vacías (por borrar sus consumos) no deben contar como actividad."""

    def test_consultar_cuenta_no_la_crea(self, client, auth_headers):
        """El GET devuelve una cuenta virtual con id nulo: nada se persistió."""
        vacio = _crear_cliente(client, "Solo Mira")
        c = _cuenta(client, vacio["id"])
        assert c["id"] is None
        assert c["transacciones"] == []
        assert c["estado"] == "abierta"
        # Consultarla de nuevo sigue sin crearla
        assert _cuenta(client, vacio["id"])["id"] is None

    def test_agregar_consumo_materializa_la_cuenta(self, client, auth_headers):
        """La cuenta nace con el primer consumo, no antes."""
        cliente = _crear_cliente(client, "Compra Luego")
        prod = _crear_producto(client, "Producto Materializa", "1000")
        assert _cuenta(client, cliente["id"])["id"] is None

        _agregar_item(client, cliente["id"], prod["id"], cantidad=1)
        assert _cuenta(client, cliente["id"])["id"] is not None

    def test_cuenta_vacia_no_cuenta_como_cliente_activo(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        con_consumo = _crear_cliente(client, "Compra De Verdad")
        prod = _crear_producto(client, "Producto Fantasma", "1000")
        _agregar_item(client, con_consumo["id"], prod["id"], cantidad=2)

        # Cuentas que quedaron vacías tras borrarles el único consumo
        for i in range(5):
            vacio = _crear_cliente(client, f"Solo Mira {i}")
            t = _agregar_item(client, vacio["id"], prod["id"], cantidad=1)
            assert client.delete(f"/api/v1/cuentas/transaccion/{t['id']}").status_code == 204

        r = _balance(client, mes, anio)["resumen"]
        assert r["clientes_activos"]["actual"] == 1
        assert r["cuentas_abiertas"] == 1
        assert r["ticket_promedio"]["actual"] == 2000  # no 2000/6

    def test_cuentas_vacias_no_generan_deudores(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        vacio = _crear_cliente(client, "Cliente Sin Consumo")
        prod = _crear_producto(client, "Producto Deudor Vacio", "1000")
        t = _agregar_item(client, vacio["id"], prod["id"], cantidad=1)
        assert client.delete(f"/api/v1/cuentas/transaccion/{t['id']}").status_code == 204

        b = _balance(client, mes, anio)
        assert all(d["nombre"] != "Cliente Sin Consumo" for d in b["deudores"])
        assert b["resumen"]["clientes_con_deuda"] == 0


class TestArrastreDeDeuda:
    """El pago parcial genera 'Traspaso de deuda' (negativo, mes actual) y
    'Deuda anterior' (positivo, mes siguiente)."""

    def test_pago_parcial_cobrado_refleja_lo_pagado(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Parcial")
        prod = _crear_producto(client, "Producto Parcial", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=10)  # 10.000
        _cerrar(client, _cuenta(client, cliente["id"])["id"], monto="4000")

        r = _balance(client, mes, anio)["resumen"]
        assert r["cobrado"]["actual"] == 4000       # incluye el traspaso negativo
        assert r["ventas"]["actual"] == 10000       # la venta real no cambia
        assert r["deuda_traspasada"] == 6000

    def test_deuda_anterior_suma_a_por_cobrar_pero_no_a_ventas(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Rollover")
        prod = _crear_producto(client, "Producto Rollover", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=10)
        _cerrar(client, _cuenta(client, cliente["id"])["id"], monto="4000")

        sig_mes = mes + 1 if mes < 12 else 1
        sig_anio = anio if mes < 12 else anio + 1
        b = _balance(client, sig_mes, sig_anio)
        assert b["resumen"]["deuda_arrastrada"] == 6000
        assert b["resumen"]["por_cobrar"]["actual"] == 6000
        assert b["resumen"]["ventas"]["actual"] == 0

    def test_productos_de_arrastre_no_contaminan_listas(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Contaminacion")
        prod = _crear_producto(client, "Producto Limpio", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=5)
        _cerrar(client, _cuenta(client, cliente["id"])["id"], monto="1000")

        sig_mes = mes + 1 if mes < 12 else 1
        sig_anio = anio if mes < 12 else anio + 1
        prohibidos = {"Deuda anterior", "Traspaso de deuda"}
        for m, a in [(mes, anio), (sig_mes, sig_anio)]:
            b = _balance(client, m, a)
            assert prohibidos.isdisjoint({p["nombre"] for p in b["productos_top"]})
            assert prohibidos.isdisjoint({p["nombre"] for p in b["stock_bajo"]})
            assert sum(v["monto"] for v in b["ventas_por_dia"]) >= 0

    def test_lineas_de_arrastre_vienen_marcadas_y_fechadas_en_su_mes(self, client, auth_headers):
        """La interfaz distingue el saldo del consumo por la marca `es_arrastre`,
        no por el nombre del producto. Y cada línea se fecha dentro del mes de su
        propia cuenta: cerrar en agosto una cuenta de mayo no debe dejar la línea
        de mayo fechada en agosto."""
        cliente = _crear_cliente(client, "Cliente Marcado")
        prod = _crear_producto(client, "Producto Marcado", "1000")
        # Cuenta explícita de marzo 2025, cerrada "hoy" con pago parcial
        _agregar_item(client, cliente["id"], prod["id"], cantidad=10, mes=3, anio=2025)
        _cerrar(client, _cuenta(client, cliente["id"], 3, 2025)["id"], monto="4000")

        # Mes cerrado: el traspaso negativo queda el último día de marzo
        pagadas = _cuenta(client, cliente["id"], 3, 2025)["transacciones_pagadas"]
        traspaso = [t for t in pagadas if t["es_arrastre"]]
        assert len(traspaso) == 1
        assert traspaso[0]["fecha_hora"].startswith("2025-03-31")
        assert all(not t["es_arrastre"] for t in pagadas if t["producto_nombre"] == "Producto Marcado")

        # Mes siguiente: la deuda anterior queda el día 1 de abril
        abril = _cuenta(client, cliente["id"], 4, 2025)
        arrastre = [t for t in abril["transacciones"] if t["es_arrastre"]]
        assert len(arrastre) == 1
        assert arrastre[0]["fecha_hora"].startswith("2025-04-01")


class TestMultiplesCuentas:
    def test_cliente_con_cuenta_pagada_y_abierta_en_el_mismo_mes(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Dos Cuentas")
        prod = _crear_producto(client, "Producto Dos Cuentas", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=3)      # 3000
        _cerrar(client, _cuenta(client, cliente["id"])["id"])             # paga
        _agregar_item(client, cliente["id"], prod["id"], cantidad=2)      # 2000 en cuenta nueva

        b = _balance(client, mes, anio)
        r = b["resumen"]
        assert r["cobrado"]["actual"] == 3000
        assert r["por_cobrar"]["actual"] == 2000
        assert r["ventas"]["actual"] == 5000
        assert r["clientes_activos"]["actual"] == 1
        assert r["cuentas_pagadas"] == 1 and r["cuentas_abiertas"] == 1
        assert len([c for c in b["clientes_top"] if c["nombre"] == "Cliente Dos Cuentas"]) == 1
        assert len([d for d in b["deudores"] if d["nombre"] == "Cliente Dos Cuentas"]) == 1


class TestDeudores:
    def test_deuda_mes_vs_deuda_total(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        ant_mes = mes - 1 if mes > 1 else 12
        ant_anio = anio if mes > 1 else anio - 1

        cliente = _crear_cliente(client, "Cliente Historico")
        prod = _crear_producto(client, "Producto Historico", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=4, mes=ant_mes, anio=ant_anio)
        _agregar_item(client, cliente["id"], prod["id"], cantidad=3)

        b = _balance(client, mes, anio)
        fila = next(d for d in b["deudores"] if d["nombre"] == "Cliente Historico")
        assert fila["deuda_mes"] == 3000
        assert fila["deuda_total"] == 7000

        # Coincide con la property Cliente.deuda que expone /clientes/
        en_listado = next(c for c in client.get("/api/v1/clientes/").json()
                          if c["nombre"] == "Cliente Historico")
        assert float(en_listado["deuda"]) == fila["deuda_total"]

    def test_clientes_con_deuda_cuenta_clientes_no_cuentas(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Multi Deuda")
        prod = _crear_producto(client, "Producto Multi Deuda", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=2)
        _cerrar(client, _cuenta(client, cliente["id"])["id"], monto="500")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=1)

        assert _balance(client, mes, anio)["resumen"]["clientes_con_deuda"] == 1


class TestComparativaYEvolucion:
    def test_variacion_pct_contra_mes_anterior(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        ant_mes = mes - 1 if mes > 1 else 12
        ant_anio = anio if mes > 1 else anio - 1
        cliente = _crear_cliente(client, "Cliente Variacion")
        prod = _crear_producto(client, "Producto Variacion", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=10, mes=ant_mes, anio=ant_anio)
        _agregar_item(client, cliente["id"], prod["id"], cantidad=15)

        r = _balance(client, mes, anio)["resumen"]
        assert r["ventas"]["anterior"] == 10000
        assert r["ventas"]["variacion_pct"] == pytest.approx(50.0)

    def test_evolucion_devuelve_n_puntos_en_orden(self, client, auth_headers):
        resp = client.get("/api/v1/balances/evolucion", params={"mes": 6, "anio": 2026, "meses": 3})
        assert resp.status_code == 200
        puntos = resp.json()["puntos"]
        assert len(puntos) == 3
        assert [(p["mes"], p["anio"]) for p in puntos] == [(4, 2026), (5, 2026), (6, 2026)]

    def test_evolucion_cruza_el_cambio_de_anio(self, client, auth_headers):
        resp = client.get("/api/v1/balances/evolucion", params={"mes": 1, "anio": 2027, "meses": 3})
        puntos = resp.json()["puntos"]
        assert [(p["mes"], p["anio"]) for p in puntos] == [(11, 2026), (12, 2026), (1, 2027)]

    @pytest.mark.parametrize("meses", [1, 25])
    def test_evolucion_meses_fuera_de_rango_422(self, client, auth_headers, meses):
        resp = client.get("/api/v1/balances/evolucion",
                          params={"mes": 6, "anio": 2026, "meses": meses})
        assert resp.status_code == 422


class TestMermasYStock:
    def test_valor_mermas_y_agrupaciones(self, client, auth_headers):
        prod = _crear_producto(client, "Producto Merma", "2000", stock=50)
        client.post("/api/v1/perdidas/", json={"producto_id": prod["id"], "cantidad": 3,
                                               "motivo": "Vencido"})
        client.post("/api/v1/perdidas/", json={"producto_id": prod["id"], "cantidad": 1})

        mes, anio = _mes_actual(client)
        b = _balance(client, mes, anio)
        assert b["resumen"]["valor_mermas"]["actual"] == 8000  # 4 unidades x 2000
        motivos = {m["etiqueta"]: m for m in b["mermas_por_motivo"]}
        assert motivos["Vencido"]["valor"] == 6000
        assert motivos["Sin motivo"]["valor"] == 2000
        assert b["mermas_por_producto"][0]["etiqueta"] == "Producto Merma"

    def test_stock_bajo_filtra_por_umbral_y_estado(self, client, auth_headers):
        _crear_producto(client, "Producto Poco Stock", "1000", stock=3)
        _crear_producto(client, "Producto Stock Sano", "1000", stock=50)
        mes, anio = _mes_actual(client)
        nombres = {p["nombre"] for p in _balance(client, mes, anio)["stock_bajo"]}
        assert "Producto Poco Stock" in nombres
        assert "Producto Stock Sano" not in nombres


class TestVentasPorDiaYCategoria:
    def test_ventas_por_dia_suman_el_total(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Dias")
        prod = _crear_producto(client, "Producto Dias", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=2, mes=mes, anio=anio, dia=5)
        _agregar_item(client, cliente["id"], prod["id"], cantidad=3, mes=mes, anio=anio, dia=12)

        b = _balance(client, mes, anio)
        por_dia = {v["dia"]: v["monto"] for v in b["ventas_por_dia"]}
        assert por_dia[5] == 2000 and por_dia[12] == 3000
        assert sum(v["monto"] for v in b["ventas_por_dia"]) == b["resumen"]["ventas"]["actual"]

    def test_producto_sin_categoria_y_pedido_personalizado(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Categoria")
        resp = client.post(f"/api/v1/cuentas/cliente/{cliente['id']}/pedido_personalizado",
                           json={"nombre": "Pedido Suelto", "precio": "2500", "cantidad": 2})
        assert resp.status_code == 200

        b = _balance(client, mes, anio)
        assert b["resumen"]["ventas"]["actual"] == 5000
        assert [c["nombre"] for c in b["ventas_por_categoria"]] == ["Sin categoría"]


class TestExportacionExcel:
    def test_export_devuelve_xlsx_con_nombre(self, client, auth_headers):
        resp = client.get("/api/v1/balances/exportar", params={"mes": 6, "anio": 2026})
        assert resp.status_code == 200
        assert "spreadsheetml.sheet" in resp.headers["content-type"]
        assert 'filename="Balance_2026-06_Junio.xlsx"' in resp.headers["content-disposition"]

    def test_export_contiene_todas_las_hojas(self, client, auth_headers):
        resp = client.get("/api/v1/balances/exportar", params={"mes": 6, "anio": 2026})
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == HOJAS_ESPERADAS

    def test_resumen_del_excel_coincide_con_el_json(self, client, auth_headers):
        mes, anio = _mes_actual(client)
        cliente = _crear_cliente(client, "Cliente Excel")
        prod = _crear_producto(client, "Producto Excel", "1000")
        _agregar_item(client, cliente["id"], prod["id"], cantidad=7)
        _cerrar(client, _cuenta(client, cliente["id"])["id"])

        esperado = _balance(client, mes, anio)["resumen"]["cobrado"]["actual"]
        resp = client.get("/api/v1/balances/exportar", params={"mes": mes, "anio": anio})
        ws = openpyxl.load_workbook(io.BytesIO(resp.content))["Resumen"]
        valores = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                   for r in range(1, ws.max_row + 1)}
        assert valores["Cobrado"] == esperado

    def test_export_de_mes_vacio_es_valido(self, client, auth_headers):
        resp = client.get("/api/v1/balances/exportar", params={"mes": 2, "anio": 2024})
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == HOJAS_ESPERADAS
        assert wb["Deudores"].max_row >= 1


class TestAutorizacion:
    def _crear_staff(self, client):
        resp = client.post("/api/v1/auth/register",
                           json={"username": "staff_balances", "password": "staffpass123", "rol": "staff"})
        assert resp.status_code in (200, 201)

    @pytest.mark.parametrize("ruta", ["/api/v1/balances/", "/api/v1/balances/evolucion",
                                      "/api/v1/balances/exportar"])
    def test_staff_no_puede_ver_balances(self, client, auth_headers, ruta):
        self._crear_staff(client)
        client.post("/api/v1/auth/logout")
        client.post("/api/v1/auth/login", data={"username": "staff_balances", "password": "staffpass123"})
        resp = client.get(ruta, params={"mes": 6, "anio": 2026})
        assert resp.status_code == 403

    @pytest.mark.parametrize("ruta", ["/api/v1/balances/", "/api/v1/balances/evolucion",
                                      "/api/v1/balances/exportar"])
    def test_sin_sesion_401(self, client, ruta):
        resp = client.get(ruta, params={"mes": 6, "anio": 2026})
        assert resp.status_code == 401
